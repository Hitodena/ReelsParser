"""Instagram Parsing API router for parsing reels from Instagram profiles."""

from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from openpyxl.utils import get_column_letter

from app.db.dao import InstagramAccountDAO
from app.exceptions import (
    AuthUnexpectedError,
    UserNotFoundError,
    UserPrivateError,
)
from app.models import InstagramAuth
from app.services import (
    BrowserManager,
    DatabaseSessionManager,
    InstagramOrchestrator,
    ProxyManager,
)

from app.api.deps import get_browser, get_db, get_orchestrator, get_proxy_manager
from .schemas import ParseReelsSchema

parsing_router = APIRouter(prefix="/instagram", tags=["Instagram"])


@parsing_router.post(
    "/parse/xlsx",
    status_code=status.HTTP_200_OK,
    summary="Parse Instagram Reels to XLSX",
    description=(
        "Parse Reels from a target Instagram profile and return an XLSX file. "
        "Uses stored account credentials to authenticate and fetch data."
    ),
    response_model=None,
    responses={
        200: {
            "description": "XLSX file with parsed reels",
            "content": {
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": {
                    "schema": {"type": "string", "format": "binary"}
                }
            },
        },
        403: {"description": "Account is private"},
        404: {
            "description": "Account not found or no valid accounts available"
        },
        500: {"description": "Internal Server Error"},
    },
)
async def parse_reels_xlsx(
    data: ParseReelsSchema,
    browser: BrowserManager = Depends(get_browser),
    db: DatabaseSessionManager = Depends(get_db),
    orchestrator: InstagramOrchestrator = Depends(get_orchestrator),
    proxy_manager: ProxyManager = Depends(get_proxy_manager),
) -> StreamingResponse:
    """
    Parse Instagram Reels and return as XLSX file.

    **Workflow:**
    1. Retrieve least-used valid account from database
    2. Get least-used proxy for request
    3. Extract fresh credentials (doc_id, variables, headers)
    4. Fetch all reels with pagination
    5. Generate XLSX file with results
    6. Update account cookies and last_used_at

    **XLSX Columns:**
    - Ссылка: Direct link to the reel
    - Просмотры: Number of views
    - Лайки: Number of likes
    - Комменты: Number of comments
    - Вирусность: Engagement Rate = (likes + comments) / views

    Args:
        data: Parsing parameters with target_username and max_reels.
        browser: Browser manager for Playwright context.
        db: Database session manager dependency.
        orchestrator: Instagram orchestrator for parsing.
        proxy_manager: Proxy manager for getting proxies.

    Returns:
        StreamingResponse: XLSX file as download attachment.

    Raises:
        HTTPException: 403 if target account is private.
        HTTPException: 404 if account not found or no valid accounts.
        HTTPException: 500 for parsing errors.

    Example:
        POST /api/instagram/parse/xlsx
        Body: {
            "target_username": "instagram",
            "max_reels": 50
        }
        Response: Binary XLSX file download
    """
    import io

    import pandas as pd

    async with db.session() as session:
        account = await InstagramAccountDAO.get_least_used(session)

    if not account:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No valid Instagram accounts available. Please add an account first.",
        )

    auth = InstagramAuth(
        login=account.login,
        password=account.password,
        cookies=account.cookies,
    )

    proxy_formatted = None
    proxy = await proxy_manager.get_least_used()
    if proxy:
        proxy_formatted = proxy.to_playwright_proxy()

    try:
        async with browser.context(proxy=proxy_formatted) as (page, ctx):
            reels, credentials = await orchestrator.full_workflow(
                page, ctx, auth, data.target_username, data.max_reels
            )
    except AuthUnexpectedError as exc:
        async with db.session() as session:
            await InstagramAccountDAO.update_by_login(
                session, auth.login, valid=False
            )
        raise HTTPException(
            status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to use account '{account.login}' - account invalidated: {exc}",
        )
    except UserPrivateError:
        raise HTTPException(
            403, f"Account '{data.target_username}' is private"
        )
    except UserNotFoundError:
        raise HTTPException(404, f"Account '{data.target_username}' not found")
    except Exception as exc:
        async with db.session() as session:
            await InstagramAccountDAO.update_by_login(
                session, auth.login, valid=False
            )
        raise HTTPException(
            status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to parse reels: {exc}",
        )

    # Update cookies and last_used_at
    async with db.session() as session:
        await InstagramAccountDAO.update_by_login(
            session,
            auth.login,
            cookies=credentials.get("cookies"),
            last_used_at=datetime.now(),
        )

    # Filter valid reels
    required_keys = ["url", "views", "likes", "comments", "virality"]
    reels = [
        r
        for r in reels
        if isinstance(r, dict) and all(k in r for k in required_keys)
    ]

    if not reels:
        raise HTTPException(500, "Failed to get reels - no valid data")

    # Create DataFrame with Russian column names
    df = pd.DataFrame(reels)
    df = df[required_keys].rename(
        columns={
            "url": "Ссылка",
            "views": "Просмотры",
            "likes": "Лайки",
            "comments": "Комменты",
            "virality": "Вирусность",
        }
    )
    df["Просмотры"] = df["Просмотры"].astype(int)
    df["Лайки"] = df["Лайки"].astype(int)
    df["Комменты"] = df["Комменты"].astype(int)
    df["Вирусность"] = df["Вирусность"].astype(float)

    # Generate XLSX with styling
    from openpyxl.styles import Alignment, Font, PatternFill

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Reels")

        # Style the worksheet
        worksheet = writer.sheets["Reels"]

        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Auto-fit column widths based on content

        for col_num, column in enumerate(df.columns, 1):
            max_length = len(str(column))  # Start with header length
            for row_num in range(2, len(df) + 2):
                cell_value = worksheet.cell(row=row_num, column=col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            # Add padding and cap at 50 characters
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[
                get_column_letter(col_num)
            ].width = adjusted_width

        # Number format for virality (percentage)
        for row_num in range(2, len(df) + 2):
            cell = worksheet.cell(row=row_num, column=5)
            cell.number_format = "0.00%"

        # Freeze header row
        worksheet.freeze_panes = "A2"

    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={data.target_username}_reels.xlsx"
        },
    )
